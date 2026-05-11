"""Excel report generation with styled sheets"""
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.models import ExtractionResult


def create_excel_report(results: List[ExtractionResult], output_path: Path) -> None:
    """
    Create Excel workbook with extraction results.

    Creates two sheets:
    - "Dados": All successfully extracted data with styled headers
    - "Resumo": Summary statistics about the extraction process

    Args:
        results: List of extraction results (successful and failed)
        output_path: Path where the Excel file should be saved

    Raises:
        IOError: If file cannot be written
    """
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create sheets
    _create_data_sheet(wb, results)
    _create_summary_sheet(wb, results)

    # Save workbook
    wb.save(output_path)


def _create_data_sheet(wb: Workbook, results: List[ExtractionResult]) -> None:
    """
    Create the "Dados" sheet with extracted form data.

    Features:
    - Blue header row with white bold text
    - Auto-sized columns
    - Frozen header row
    - Only includes successful extractions

    Args:
        wb: Workbook to add sheet to
        results: Extraction results to include
    """
    ws = wb.create_sheet("Dados")

    # Get successful results with data
    successful = [r for r in results if r.success and r.data]

    if not successful:
        # Empty sheet if no successful extractions
        ws.append(["No successful extractions"])
        return

    # Get headers from first successful result
    headers = list(successful[0].data.to_dict().keys())

    # Write headers with styling
    ws.append(headers)

    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for result in successful:
        row_data = list(result.data.to_dict().values())
        ws.append(row_data)

    # Auto-size columns
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = len(headers[col_num - 1])

        # Check data rows for max length
        for row_num in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width (add padding)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"


def _create_summary_sheet(wb: Workbook, results: List[ExtractionResult]) -> None:
    """
    Create the "Resumo" sheet with extraction statistics.

    Shows:
    - Total files processed
    - Successful extractions
    - Failed extractions
    - Success rate percentage

    Args:
        wb: Workbook to add sheet to
        results: Extraction results to summarize
    """
    ws = wb.create_sheet("Resumo")

    # Calculate statistics
    total = len(results)
    successful = sum(1 for r in results if r.success)
    failed = total - successful
    success_rate = (successful / total * 100) if total > 0 else 0.0

    # Add title
    ws.append(["Estatísticas da Extração"])
    ws["A1"].font = Font(bold=True, size=14)

    # Add blank row
    ws.append([])

    # Add statistics
    ws.append(["Total de arquivos processados:", total])
    ws.append(["Extrações bem-sucedidas:", successful])
    ws.append(["Extrações com falha:", failed])
    ws.append(["Taxa de sucesso:", f"{success_rate:.1f}%"])

    # Style labels
    for row in range(3, 7):
        ws[f"A{row}"].font = Font(bold=True)

    # Auto-size columns
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
