"""
Extração em lote de PDFs de formulários do Projeto Ancestralidades
Uso: python extrair_pdfs.py ./pasta_com_pdfs
"""

import anthropic
import base64
import json
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Configuração ──────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Você é um extrator de dados de formulários manuscritos de pesquisa.
Extraia EXATAMENTE os campos abaixo do formulário e retorne APENAS um JSON válido,
sem markdown, sem texto adicional.

Campos obrigatórios:
{
  "codigo_participante": "",
  "data_coleta": "",
  "nome_completo": "",
  "local_nascimento_criacao": "",
  "sexo_biologico": "",
  "genero": "",
  "idade": "",
  "data_nascimento": "",
  "cor_raca": "",
  "nome_mae": "",
  "cor_raca_mae": "",
  "local_nascimento_mae": "",
  "nome_pai": "",
  "cor_raca_pai": "",
  "local_nascimento_pai": "",
  "medicamentos": "",
  "troca_tratamento_6meses": "",
  "interrompeu_medicacao": "",
  "reacao_adversa": "",
  "ultima_pressao_arterial": "",
  "ultimo_exame_glicemia": "",
  "ultimo_exame_colesterol": "",
  "diagnosticos_proprios": "",
  "historico_familiar": "",
  "altura_cm": "",
  "peso_kg": "",
  "imc": "",
  "gordura_corporal_pct": "",
  "massa_muscular_pct": "",
  "gordura_visceral": "",
  "cintura_cm": "",
  "barriga_cm": "",
  "quadril_cm": "",
  "frequencia_cardiaca_bpm": "",
  "glicemia_mgdl": "",
  "temperatura_corporal": "",
  "pressao_arterial_pas": "",
  "pressao_arterial_pad": "",
  "horas_sem_alimentar": ""
}

Regras:
- Para campos de múltipla escolha, liste os itens marcados separados por " | "
- Se o campo estiver em branco ou ilegível, use ""
- Retorne SOMENTE o JSON, nada mais"""

COLUNAS = [
    "arquivo",
    "codigo_participante",
    "data_coleta",
    "nome_completo",
    "local_nascimento_criacao",
    "sexo_biologico",
    "genero",
    "idade",
    "data_nascimento",
    "cor_raca",
    "nome_mae",
    "cor_raca_mae",
    "local_nascimento_mae",
    "nome_pai",
    "cor_raca_pai",
    "local_nascimento_pai",
    "medicamentos",
    "troca_tratamento_6meses",
    "interrompeu_medicacao",
    "reacao_adversa",
    "ultima_pressao_arterial",
    "ultimo_exame_glicemia",
    "ultimo_exame_colesterol",
    "diagnosticos_proprios",
    "historico_familiar",
    "altura_cm",
    "peso_kg",
    "imc",
    "gordura_corporal_pct",
    "massa_muscular_pct",
    "gordura_visceral",
    "cintura_cm",
    "barriga_cm",
    "quadril_cm",
    "frequencia_cardiaca_bpm",
    "glicemia_mgdl",
    "temperatura_corporal",
    "pressao_arterial_pas",
    "pressao_arterial_pad",
    "horas_sem_alimentar",
]

CABECALHOS_PT = {
    "arquivo": "Arquivo",
    "codigo_participante": "Código Participante",
    "data_coleta": "Data Coleta",
    "nome_completo": "Nome Completo",
    "local_nascimento_criacao": "Local Nascimento/Criação",
    "sexo_biologico": "Sexo Biológico",
    "genero": "Gênero",
    "idade": "Idade",
    "data_nascimento": "Data de Nascimento",
    "cor_raca": "Cor/Raça",
    "nome_mae": "Nome da Mãe",
    "cor_raca_mae": "Cor/Raça da Mãe",
    "local_nascimento_mae": "Local Nasc. Mãe",
    "nome_pai": "Nome do Pai",
    "cor_raca_pai": "Cor/Raça do Pai",
    "local_nascimento_pai": "Local Nasc. Pai",
    "medicamentos": "Medicamentos",
    "troca_tratamento_6meses": "Troca Tratamento (6m)",
    "interrompeu_medicacao": "Interrompeu Medicação",
    "reacao_adversa": "Reação Adversa",
    "ultima_pressao_arterial": "Última Pressão Arterial",
    "ultimo_exame_glicemia": "Último Exame Glicemia",
    "ultimo_exame_colesterol": "Último Exame Colesterol",
    "diagnosticos_proprios": "Diagnósticos",
    "historico_familiar": "Histórico Familiar",
    "altura_cm": "Altura (cm)",
    "peso_kg": "Peso (kg)",
    "imc": "IMC",
    "gordura_corporal_pct": "Gordura Corporal (%)",
    "massa_muscular_pct": "Massa Muscular (%)",
    "gordura_visceral": "Gordura Visceral",
    "cintura_cm": "Cintura (cm)",
    "barriga_cm": "Barriga (cm)",
    "quadril_cm": "Quadril (cm)",
    "frequencia_cardiaca_bpm": "FC (bpm)",
    "glicemia_mgdl": "Glicemia (mg/dL)",
    "temperatura_corporal": "Temperatura (°C)",
    "pressao_arterial_pas": "PAS (mmHg)",
    "pressao_arterial_pad": "PAD (mmHg)",
    "horas_sem_alimentar": "Horas sem Alimentar",
}

# ── Funções ───────────────────────────────────────────────────────────────────


def pdf_para_base64(caminho: Path) -> str:
    return base64.standard_b64encode(caminho.read_bytes()).decode("utf-8")


def extrair_dados(cliente: anthropic.Anthropic, caminho_pdf: Path) -> dict:
    dados_b64 = pdf_para_base64(caminho_pdf)

    resposta = cliente.messages.create(
        model="claude-opus-4-5",
        max_tokens=2000,
        system=SYSTEM_PROMPT,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": dados_b64,
                        },
                    },
                    {"type": "text", "text": "Extraia os dados deste formulário."},
                ],
            }
        ],
    )

    texto = resposta.content[0].text.strip()
    # Remove possíveis blocos markdown
    if texto.startswith("```"):
        texto = texto.split("```")[1]
        if texto.startswith("json"):
            texto = texto[4:]
    return json.loads(texto)


def criar_excel(registros: list[dict], caminho_saida: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados"

    # Estilo do cabeçalho
    cor_cabecalho = "1F4E79"
    fonte_cabecalho = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    fill_cabecalho = PatternFill("solid", start_color=cor_cabecalho)
    alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Escreve cabeçalhos
    for col_idx, campo in enumerate(COLUNAS, start=1):
        celula = ws.cell(row=1, column=col_idx, value=CABECALHOS_PT.get(campo, campo))
        celula.font = fonte_cabecalho
        celula.fill = fill_cabecalho
        celula.alignment = alinhamento_centro

    ws.row_dimensions[1].height = 35

    # Cores alternadas para linhas
    fill_par = PatternFill("solid", start_color="DEEAF1")
    fill_impar = PatternFill("solid", start_color="FFFFFF")
    fonte_dados = Font(name="Arial", size=10)

    for linha_idx, registro in enumerate(registros, start=2):
        fill = fill_par if linha_idx % 2 == 0 else fill_impar
        for col_idx, campo in enumerate(COLUNAS, start=1):
            valor = registro.get(campo, "")
            celula = ws.cell(row=linha_idx, column=col_idx, value=valor)
            celula.font = fonte_dados
            celula.fill = fill
            celula.alignment = Alignment(vertical="top", wrap_text=True)

    # Largura das colunas
    larguras = {
        "arquivo": 15,
        "nome_completo": 30,
        "medicamentos": 50,
        "diagnosticos_proprios": 45,
        "historico_familiar": 45,
        "local_nascimento_criacao": 30,
    }
    largura_padrao = 18

    for col_idx, campo in enumerate(COLUNAS, start=1):
        letra = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letra].width = larguras.get(campo, largura_padrao)

    # Congela a primeira linha
    ws.freeze_panes = "A2"

    # Aba de resumo
    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo["A1"] = "Total de participantes"
    ws_resumo["B1"] = f"=COUNTA(Dados!A2:A{1 + len(registros)})"
    ws_resumo["A1"].font = Font(bold=True, name="Arial")
    ws_resumo["B1"].font = Font(name="Arial")

    wb.save(caminho_saida)


# ── Main ──────────────────────────────────────────────────────────────────────


def main():
    pasta = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(".")
    pdfs = sorted(pasta.glob("*.pdf"))

    if not pdfs:
        print(f"Nenhum PDF encontrado em: {pasta.resolve()}")
        sys.exit(1)

    print(f"📂 {len(pdfs)} PDF(s) encontrado(s) em {pasta.resolve()}\n")

    cliente = anthropic.Anthropic()  # Usa ANTHROPIC_API_KEY do ambiente
    registros = []
    erros = []

    for i, pdf in enumerate(pdfs, start=1):
        print(f"[{i}/{len(pdfs)}] Processando: {pdf.name} ...", end=" ", flush=True)
        try:
            dados = extrair_dados(cliente, pdf)
            dados["arquivo"] = pdf.name
            registros.append(dados)
            print("✅")
        except Exception as e:
            print(f"❌ Erro: {e}")
            erros.append({"arquivo": pdf.name, "erro": str(e)})
            registros.append({"arquivo": pdf.name})

        # Pequena pausa para respeitar rate limits
        if i < len(pdfs):
            time.sleep(1)

    saida = pasta / "resultados_ancestralidades.xlsx"
    criar_excel(registros, saida)

    print(f"\n✅ Planilha salva em: {saida}")
    print(f"   {len(registros) - len(erros)} participante(s) extraído(s) com sucesso")
    if erros:
        print(f"   ⚠️  {len(erros)} erro(s):")
        for e in erros:
            print(f"      - {e['arquivo']}: {e['erro']}")


if __name__ == "__main__":
    main()