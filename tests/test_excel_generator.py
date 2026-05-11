"""Tests for Excel report generation"""
import pytest
from pathlib import Path
from openpyxl import load_workbook

from src.models import FormData, ExtractionResult
from src.excel_generator import create_excel_report


@pytest.fixture
def sample_results():
    """Sample extraction results for testing"""
    data1 = FormData(
        quilombo="Quilombo A",
        arquivo="test1.pdf",
        codigo_participante="001",
        data_coleta="01/01/2024",
        nome_completo="João Silva",
        local_nascimento_criacao="Porto Alegre",
        sexo_biologico="Masculino",
        genero="Masculino",
        idade="45",
        data_nascimento="01/01/1979",
        cor_raca="Pardo",
        nome_mae="Maria Silva",
        cor_raca_mae="Parda",
        local_nascimento_mae="Porto Alegre",
        nome_pai="José Silva",
        cor_raca_pai="Pardo",
        local_nascimento_pai="Porto Alegre",
        medicamentos="Losartana 50mg",
        troca_tratamento_6meses="Não",
        interrompeu_medicacao="Não",
        reacao_adversa="Não",
        ultima_pressao_arterial="120/80",
        ultimo_exame_glicemia="95",
        ultimo_exame_colesterol="180",
        diagnosticos_proprios="Hipertensão",
        historico_familiar="Diabetes",
        altura_cm="175",
        peso_kg="80",
        imc="26.1",
        gordura_corporal_pct="22",
        massa_muscular_pct="40",
        gordura_visceral="8",
        cintura_cm="90",
        barriga_cm="92",
        quadril_cm="95",
        frequencia_cardiaca_bpm="72",
        glicemia_mgdl="95",
        temperatura_corporal="36.5",
        pressao_arterial_pas="120",
        pressao_arterial_pad="80",
        horas_sem_alimentar="8"
    )

    result1 = ExtractionResult(
        filename="test1.pdf",
        success=True,
        data=data1,
        error=None,
        attempts=1,
        duration_seconds=2.5
    )

    data2 = FormData(
        quilombo="Quilombo B",
        arquivo="test2.pdf",
        codigo_participante="002",
        data_coleta="02/01/2024",
        nome_completo="Maria Santos",
        local_nascimento_criacao="Canoas",
        sexo_biologico="Feminino",
        genero="Feminino",
        idade="38",
        data_nascimento="02/02/1986",
        cor_raca="Negra",
        nome_mae="Ana Santos",
        cor_raca_mae="Negra",
        local_nascimento_mae="Canoas",
        nome_pai="Pedro Santos",
        cor_raca_pai="Negro",
        local_nascimento_pai="Canoas",
        medicamentos="",
        troca_tratamento_6meses="",
        interrompeu_medicacao="",
        reacao_adversa="",
        ultima_pressao_arterial="110/70",
        ultimo_exame_glicemia="88",
        ultimo_exame_colesterol="170",
        diagnosticos_proprios="",
        historico_familiar="Hipertensão",
        altura_cm="160",
        peso_kg="65",
        imc="25.4",
        gordura_corporal_pct="28",
        massa_muscular_pct="35",
        gordura_visceral="6",
        cintura_cm="85",
        barriga_cm="87",
        quadril_cm="92",
        frequencia_cardiaca_bpm="68",
        glicemia_mgdl="88",
        temperatura_corporal="36.7",
        pressao_arterial_pas="110",
        pressao_arterial_pad="70",
        horas_sem_alimentar="10"
    )

    result2 = ExtractionResult(
        filename="test2.pdf",
        success=True,
        data=data2,
        error=None,
        attempts=1,
        duration_seconds=2.8
    )

    return [result1, result2]


def test_create_excel_report(tmp_path, sample_results):
    """Test successful Excel report generation with styled sheets"""
    output_file = tmp_path / "report.xlsx"

    create_excel_report(sample_results, output_file)

    # Verify file exists
    assert output_file.exists()

    # Load workbook and verify structure
    wb = load_workbook(output_file)

    # Should have two sheets
    assert len(wb.sheetnames) == 2
    assert "Dados" in wb.sheetnames
    assert "Resumo" in wb.sheetnames

    # Verify data sheet
    dados = wb["Dados"]

    # Check headers (row 1)
    assert dados["A1"].value == "quilombo"
    assert dados["B1"].value == "arquivo"
    assert dados["C1"].value == "codigo_participante"

    # Check header styling
    assert dados["A1"].font.bold is True
    assert dados["A1"].fill.start_color.rgb == "004472C4"  # Blue fill

    # Check data rows
    assert dados["A2"].value == "Quilombo A"
    assert dados["B2"].value == "test1.pdf"
    assert dados["C2"].value == "001"

    assert dados["A3"].value == "Quilombo B"
    assert dados["B3"].value == "test2.pdf"
    assert dados["C3"].value == "002"

    # Check frozen panes (header row should be frozen)
    assert dados.freeze_panes == "A2"

    # Verify summary sheet
    resumo = wb["Resumo"]

    # Check summary content
    assert resumo["A1"].value == "Estatísticas da Extração"
    assert resumo["A1"].font.bold is True
    assert resumo["A1"].font.size == 14

    assert resumo["A3"].value == "Total de arquivos processados:"
    assert resumo["B3"].value == 2

    assert resumo["A4"].value == "Extrações bem-sucedidas:"
    assert resumo["B4"].value == 2

    assert resumo["A5"].value == "Extrações com falha:"
    assert resumo["B5"].value == 0

    assert resumo["A6"].value == "Taxa de sucesso:"
    assert resumo["B6"].value == "100.0%"


def test_create_excel_with_failed_extraction(tmp_path):
    """Test report generation with failed extractions"""
    failed_result = ExtractionResult(
        filename="failed.pdf",
        success=False,
        data=None,
        error="API timeout after 3 attempts",
        attempts=3,
        duration_seconds=15.0
    )

    data = FormData(
        quilombo="Quilombo C",
        arquivo="success.pdf",
        codigo_participante="003",
        data_coleta="03/01/2024",
        nome_completo="Pedro Costa",
        local_nascimento_criacao="Pelotas",
        sexo_biologico="Masculino",
        genero="Masculino",
        idade="50",
        data_nascimento="03/03/1974",
        cor_raca="Negro",
        nome_mae="Julia Costa",
        cor_raca_mae="Negra",
        local_nascimento_mae="Pelotas",
        nome_pai="Carlos Costa",
        cor_raca_pai="Negro",
        local_nascimento_pai="Pelotas",
        medicamentos="Metformina",
        troca_tratamento_6meses="Sim",
        interrompeu_medicacao="Não",
        reacao_adversa="Não",
        ultima_pressao_arterial="130/85",
        ultimo_exame_glicemia="110",
        ultimo_exame_colesterol="200",
        diagnosticos_proprios="Diabetes tipo 2",
        historico_familiar="Diabetes",
        altura_cm="170",
        peso_kg="85",
        imc="29.4",
        gordura_corporal_pct="30",
        massa_muscular_pct="38",
        gordura_visceral="10",
        cintura_cm="95",
        barriga_cm="98",
        quadril_cm="100",
        frequencia_cardiaca_bpm="75",
        glicemia_mgdl="110",
        temperatura_corporal="36.6",
        pressao_arterial_pas="130",
        pressao_arterial_pad="85",
        horas_sem_alimentar="12"
    )

    success_result = ExtractionResult(
        filename="success.pdf",
        success=True,
        data=data,
        error=None,
        attempts=1,
        duration_seconds=2.3
    )

    results = [failed_result, success_result]
    output_file = tmp_path / "report_with_failure.xlsx"

    create_excel_report(results, output_file)

    # Verify file exists
    assert output_file.exists()

    # Load and verify summary
    wb = load_workbook(output_file)
    resumo = wb["Resumo"]

    assert resumo["B3"].value == 2  # Total
    assert resumo["B4"].value == 1  # Success
    assert resumo["B5"].value == 1  # Failed
    assert resumo["B6"].value == "50.0%"  # Success rate

    # Verify only successful extraction appears in data sheet
    dados = wb["Dados"]
    # Row 1 is header, row 2 should be the successful extraction
    assert dados["A2"].value == "Quilombo C"
    assert dados["B2"].value == "success.pdf"
    # Row 3 should be empty (failed extraction not included)
    assert dados["A3"].value is None
